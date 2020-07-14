from os.path import join

from openpyxl import load_workbook

from controladores.PadronAfip import PadronAfip


class ConsultaPadron:

    archivo_excel = join('plantillas', 'base cuit.xlsx')

    def recorre_archivo(self):
        wb = load_workbook(self.archivo_excel)
        ws = wb.active
        padron = PadronAfip()
        fila = 1
        for row in ws.iter_rows():
            print(row[0].value, row[1].value)
            if str(row[1].value).isdigit():
                if padron.ConsultarPersona(row[1].value):
                    ws.cell(row=fila, column=3).value = padron.imp_iva
                    ws.cell(row=fila, column=4).value = padron.denominacion
                    ws.cell(row=fila, column=5).value = padron.domicilio
                    ws.cell(row=fila, column=6).value = padron.localidad
                    ws.cell(row=fila, column=7).value = padron.provincia
                    ws.cell(row=fila, column=8).value = padron.cod_postal
                else:
                    ws.cell(row=fila, column=3).value = 'NO ENCONTRADO'
            fila += 1
        wb.save(self.archivo_excel)

if __name__ == '__main__':
    padron = ConsultaPadron()
    padron.recorre_archivo()

