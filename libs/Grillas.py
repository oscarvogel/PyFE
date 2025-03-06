# coding=utf-8
import datetime
import decimal
import re

import xlsxwriter
from PyQt5 import QtCore
from PyQt5.QtCore import QAbstractTableModel, Qt, QVariant
from PyQt5.QtGui import QFont, QColor
from PyQt5.QtWidgets import QTableWidget, QTableWidgetItem, QFileDialog, QAbstractItemView

from openpyxl.reader.excel import load_workbook

from libs import Ventanas
from libs.Utiles import EsVerdadero, AbrirArchivo, saveFileDialog


class Grilla(QTableWidget):

    #columnas a ocultar
    columnasOcultas = []

    #lista con las cabeceras de la grilla
    cabeceras = []

    #tabla desde la cual obtener los datos
    tabla = None

    #campos de la tabla
    campos = None

    #condiciones para filtrar los datos
    condiciones = None

    #cantidad de registros a mostrar
    limite = 100

    #columnas habilitadas
    columnasHabilitadas = []

    #campos tabla
    camposTabla = None

    #valores a cargar
    data = None

    #indica si esta en la grilla
    engrilla = False

    #indica si las columnas son seleccionables o no
    enabled = False

    #widget para las columnas
    widgetCol = {}

    #color para la columna
    backgroundColorCol = {}

    #tamaño de la fuente
    tamanio = 12

    #emit signal
    keyPressed = QtCore.pyqtSignal(int)
    
    #formatos de las columnas
    formatos = {}

    def __init__(self, *args, **kwargs):

        QTableWidget.__init__(self, *args)
        if 'tamanio' in kwargs:
            self.tamanio = kwargs['tamanio']
        font = QFont()
        font.setPointSizeF(self.tamanio)
        self.setFont(font)
        if 'habilitarorden' in kwargs:
            self.setSortingEnabled(kwargs['habilitarorden'])
        else:
            self.setSortingEnabled(True)
        # self.itemClicked.connect(self.handleItemClicked)
        self.setEditTriggers(QAbstractItemView.AllEditTriggers)#para que se pueda editar el contenido con solo un click
        
        if 'enabled' in kwargs:
            self.enabled = kwargs['enabled']
        else:
            self.enabled = True
        
        self.setEnabled(self.enabled)


    def ArmaCabeceras(self, cabeceras=None):

        if not cabeceras:
            cabeceras = self.cabeceras

        self.setColumnCount(len(cabeceras))

        for col in range(0, len(cabeceras)):
            self.setHorizontalHeaderItem(col, QTableWidgetItem(cabeceras[col]))

        self.resizeRowsToContents()
        self.resizeColumnsToContents()
        self.cabeceras = cabeceras
        self.OcultaColumnas()

    def AgregaItem(self, items=None,
                   backgroundColor=QColor(255, 255, 255), readonly=False):

        if items:
            col = 0
            cantFilas = self.rowCount() + 1
            self.setRowCount(cantFilas)
            for x in items:
                flags = QtCore.Qt.ItemIsSelectable
                if isinstance(x, bool):
                    item = QTableWidgetItem(x)
                    if x:
                        item.setCheckState(QtCore.Qt.Checked)
                    else:
                        item.setCheckState(QtCore.Qt.Unchecked)
                    self.formatos[col] = 'Bool'
                elif isinstance(x, (int, float, decimal.Decimal)):
                    item = QTableWidgetItem(str(x))
                    item.setTextAlignment(Qt.AlignRight)
                    self.formatos[col] = 'Decimal'
                # en caso de que sea formato de fecha
                elif isinstance(x, (datetime.date)):
                    fecha = x.strftime('%d/%m/%Y')
                    item = QTableWidgetItem(fecha)
                    self.formatos[col] = 'Date'
                # en caso de que sea formato de hora
                elif isinstance(x, (datetime.time)):
                    fecha = x.strftime('%H:%M:%S')
                    item = QTableWidgetItem(fecha)
                    self.formatos[col] = 'Time'
                elif isinstance(x, (bytes)):
                    if EsVerdadero(x):
                        item = 'Si'
                    else:
                        item = 'No'
                    item = QTableWidgetItem(QTableWidgetItem(x))
                    self.formatos[col] = 'Bytes'
                else:
                    item = QTableWidgetItem(QTableWidgetItem(x))
                    self.formatos[col] = 'String'

                if readonly:
                    flags = QtCore.Qt.ItemIsSelectable
                elif col in self.columnasHabilitadas:
                    if isinstance(x, (bool)):
                        flags |= QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEditable | QtCore.Qt.ItemIsEnabled
                    else:
                        flags |= QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsEditable
                else:
                    if self.enabled and not readonly:
                        flags |= QtCore.Qt.ItemIsEnabled

                item.setFlags(flags)
                if col in self.backgroundColorCol:
                    item.setBackground(self.backgroundColorCol[col])

                if backgroundColor and isinstance(backgroundColor, QColor):
                    item.setBackground(backgroundColor)

                self.setItem(cantFilas - 1, col, item)
                # if self.widgetCol:
                #     self.ArmaWidgetCol(col)
                if col in self.widgetCol:
                    widgetColumna = self.widgetCol[col]
                    self.setItemDelegateForColumn(col, widgetColumna)
                    # self.setItemDelegate(widgetColumna)
                    # self.setCellWidget(cantFilas - 1, col, widgetColumna)
                col += 1
            self.resizeRowsToContents()
            self.resizeColumnsToContents()

    def OcultaColumnas(self):
        for x in self.columnasOcultas:
            self.hideColumn(x)

    def ModificaItem(self, valor, fila, col, backgroundColor=None):
        """

        :param fila: la fila que se quiere modificar
        :param valor: valor a modificar
        :type col: entero en caso de indicar un numero de columna y string si quiero el nombre
        """
        if isinstance(valor, (int, float, decimal.Decimal)):
            item = QTableWidgetItem(str(valor))
        else:
            item = QTableWidgetItem(valor)

        if not isinstance(col, int):
            numCol = self.cabeceras.index(col)
        else:
            numCol = col

        if numCol in self.columnasHabilitadas:
            item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsEditable)
        else:
            # item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
            item.setFlags(QtCore.Qt.ItemIsSelectable)

        if col in self.backgroundColorCol:
            item.setBackground(self.backgroundColorCol[col])

        self.setItem(fila, numCol, item)
        self.resizeColumnsToContents()
        #self.dataChanged()

    def ObtenerItem(self, fila, col):

        if isinstance(col, int):
            numCol = col
        else:
            numCol = self.cabeceras.index(col)

        try:
            item = self.item(fila, numCol).text()
        except:
            item = ''
        # model = self.model()
        # index = model.index(fila, numCol)
        #
        # print(model.data(index))
        #return model.data(index).replace(',','.') if model.data(index) else ''
        return item.replace(',','.') if item else 0

    def ObtenerItemNumerico(self, fila, col):

        if isinstance(col, int):
            numCol = col
        else:
            numCol = self.cabeceras.index(col)

        try:
            item = self.item(fila, numCol)
            if item.checkState() == QtCore.Qt.Checked:
                item = True
            else:
                item = item.text()
                item = re.sub("[^0123456789\.]","",item)

            item = float(item)
        except:
            item = 0

        # return item.replace(',','.') if item else 0
        return item

    def CargaDatos(self, avance=None):

        self.blockSignals(True)
        self.setRowCount(0)
        self.blockSignals(False)

    def focusInEvent(self, *args, **kwargs):
        self.engrilla = True

    def focusOutEvent(self, *args, **kwargs):
        self.engrilla = False

    def ExportaExcel(self, columnas=None, archivo="", titulo="", nuevo=True, hoja='', fila=0, col=0):

        if not columnas:
            columnas = self.cabeceras

        if nuevo:
            archivo = archivo.replace('.', '').replace('/', '')
        if not archivo.startswith("excel"):
            archivo = "excel/" + archivo

        if nuevo:
            cArchivo = saveFileDialog(filename=archivo,
                                      files="Archivos de Excel (*.xlsx)")
        else:
            cArchivo = archivo
        # cArchivo = QFileDialog.getSaveFileName(caption="Guardar archivo", directory="", filter="*.XLSX")
        if not cArchivo:
            return

        if nuevo:
            workbook = xlsxwriter.Workbook(cArchivo)
        else:
            try:
                # Carga el archivo
                workbook = load_workbook(cArchivo)
            except:
                Ventanas.showAlert("Sistema", f"Ocurrio un error al intentar abrir el archivo {cArchivo}")
                return

        if hoja:
            # Selecciona la hoja por su nombre
            worksheet = workbook[hoja]
        else:
            #creamos una hoja nueva
            worksheet = workbook.add_worksheet()

        formato_fecha = workbook.add_format({'num_format': 'dd/mm/yyyy'})
        # fila = 0
        # col = 0
        if titulo:
            worksheet.write(fila, col, titulo)
            fila += 2

        for c in columnas:
            worksheet.write(fila, col, c)
            col += 1

        fila += 1
        for row in range(self.rowCount()):
            col = 0
            for c in columnas:
                dato = self.ObtenerItem(fila=row, col=c)
                if isinstance(dato, bool):
                    dato = 'SI' if dato else 'NO'
                else:
                    dato = dato.strip()
                try:
                    dato = float(dato)
                except:
                    if dato.isdigit():
                        dato = int(dato)
                if self.formatos[col] == 'Date':

                    worksheet.write_datetime(fila, col, datetime.datetime.strptime(dato, '%d/%m/%Y').date(),
                                             formato_fecha)
                else:
                    worksheet.write(fila, col, dato)
                col += 1
            fila += 1

        # cabeceras_excel = [{'header': x} for x in columnas]
        # worksheet.add_table(0, 0, fila, col-1, cabeceras_excel)
        workbook.close()
        AbrirArchivo(cArchivo)


    def keyPressEvent(self, event):
        super(Grilla, self).keyPressEvent(event)
        self.keyPressed.emit(event.key())

